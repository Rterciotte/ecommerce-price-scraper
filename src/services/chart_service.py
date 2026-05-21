from openpyxl.chart import (
    BarChart,
    PieChart,
    Reference
)


def create_charts(workbook, df):

    """
    Create charts inside dashboard worksheet.
    """

    dashboard = workbook["Dashboard"]

    # ======================================
    # CREATE SUPPORT DATA
    # ======================================

    rating_summary = (
        df.groupby("Rating")["Price"]
        .mean()
        .reset_index()
    )

    # Write chart data into hidden area
    start_row = 20

    dashboard.cell(
        row=start_row,
        column=1,
        value="Rating"
    )

    dashboard.cell(
        row=start_row,
        column=2,
        value="Average Price"
    )

    for index, row in rating_summary.iterrows():

        dashboard.cell(
            row=start_row + index + 1,
            column=1,
            value=row["Rating"]
        )

        dashboard.cell(
            row=start_row + index + 1,
            column=2,
            value=float(row["Price"])
        )

    # ======================================
    # BAR CHART
    # ======================================

    bar_chart = BarChart()

    bar_chart.title = "Average Price by Rating"

    bar_chart.y_axis.title = "Price"

    bar_chart.x_axis.title = "Rating"

    data = Reference(
        dashboard,
        min_col=2,
        min_row=start_row,
        max_row=start_row + len(rating_summary)
    )

    categories = Reference(
        dashboard,
        min_col=1,
        min_row=start_row + 1,
        max_row=start_row + len(rating_summary)
    )

    bar_chart.add_data(
        data,
        titles_from_data=True
    )

    bar_chart.set_categories(categories)

    bar_chart.height = 10

    bar_chart.width = 18

    dashboard.add_chart(
        bar_chart,
        "J4"
    )

    # ======================================
    # PIE CHART
    # ======================================

    stock_summary = (
        df.groupby("Availability")
        .size()
        .reset_index(name="count")
    )

    pie_start = 20

    dashboard.cell(
        row=pie_start,
        column=5,
        value="Availability"
    )

    dashboard.cell(
        row=pie_start,
        column=6,
        value="Count"
    )

    for index, row in stock_summary.iterrows():

        dashboard.cell(
            row=pie_start + index + 1,
            column=5,
            value=row["Availability"]
        )

        dashboard.cell(
            row=pie_start + index + 1,
            column=6,
            value=row["count"]
        )

    pie = PieChart()

    pie.title = "Stock Availability"

    labels = Reference(
        dashboard,
        min_col=5,
        min_row=pie_start + 1,
        max_row=pie_start + len(stock_summary)
    )

    data = Reference(
        dashboard,
        min_col=6,
        min_row=pie_start,
        max_row=pie_start + len(stock_summary)
    )

    pie.add_data(
        data,
        titles_from_data=True
    )

    pie.set_categories(labels)

    pie.height = 10

    pie.width = 14

    dashboard.add_chart(
        pie,
        "J22"
    )