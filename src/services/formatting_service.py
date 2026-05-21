from openpyxl import load_workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter


# ======================================
# COLORS & STYLES
# ======================================

HEADER_FILL = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=12
)

ALT_ROW_FILL = PatternFill(
    start_color="F4F6F8",
    end_color="F4F6F8",
    fill_type="solid"
)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)


def format_excel(file_path, logger):

    """
    Apply professional formatting to Excel report.

    Args:
        file_path (str):
            Excel file path.

        logger:
            Application logger instance.
    """

    logger.info(
        "Starting Excel formatting"
    )

    workbook = load_workbook(file_path)

    for sheet in workbook.worksheets:

        logger.info(
            f"Formatting worksheet: {sheet.title}"
        )

        # ======================================
        # FREEZE HEADER
        # ======================================

        sheet.freeze_panes = "A2"

        # ======================================
        # HEADER STYLING
        # ======================================

        for cell in sheet[1]:

            cell.fill = HEADER_FILL

            cell.font = HEADER_FONT

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = THIN_BORDER

        # ======================================
        # DATA ROW STYLING
        # ======================================

        for row in sheet.iter_rows(min_row=2):

            # Apply borders
            for cell in row:
                cell.border = THIN_BORDER

            # Zebra stripe effect
            if row[0].row % 2 == 0:

                for cell in row:
                    cell.fill = ALT_ROW_FILL

        # ======================================
        # AUTO COLUMN WIDTH
        # ======================================

        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    if cell.value:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except:
                    pass

            adjusted_width = min(max_length + 5, 40)

            sheet.column_dimensions[
                column_letter
            ].width = adjusted_width

        # ======================================
        # AUTO FILTER
        # ======================================

        sheet.auto_filter.ref = sheet.dimensions

    workbook.save(file_path)

    logger.info(
        "Excel formatting completed successfully"
    )