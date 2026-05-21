from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)


# ======================================
# COLORS & STYLES
# ======================================

TITLE_FILL = PatternFill(
    start_color="1F1F1F",
    end_color="1F1F1F",
    fill_type="solid"
)

CARD_FILL = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

CARD_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=12
)

VALUE_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=18
)

TITLE_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=22
)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)


def create_dashboard(workbook, df, logger):

    """
    Create KPI dashboard worksheet.

    Args:
        workbook:
            OpenPyXL workbook instance.

        df:
            Pandas DataFrame containing scraped data.

        logger:
            Application logger instance.
    """

    logger.info(
        "Creating dashboard worksheet"
    )

    # Remove dashboard if already exists
    if "Dashboard" in workbook.sheetnames:
        del workbook["Dashboard"]

    dashboard = workbook.create_sheet("Dashboard")

    # ======================================
    # DASHBOARD TITLE
    # ======================================

    dashboard.merge_cells("A1:H2")

    title_cell = dashboard["A1"]

    title_cell.value = "ECOMMERCE SCRAPING DASHBOARD"

    title_cell.fill = TITLE_FILL

    title_cell.font = TITLE_FONT

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Remove visual split line
    dashboard.row_dimensions[1].height = 35
    dashboard.row_dimensions[2].height = 5

    # ======================================
    # KPI METRICS
    # ======================================

    total_products = len(df)

    average_price = round(
        df["Price"].mean(),
        2
    )

    max_price = round(
        df["Price"].max(),
        2
    )

    min_price = round(
        df["Price"].min(),
        2
    )

    top_rating = (
        df["Rating"]
        .mode()[0]
    )

    metrics = [
        ("Total Products", f"{total_products}"),
        ("Average Price", f"£ {average_price}"),
        ("Highest Price", f"£ {max_price}"),
        ("Lowest Price", f"£ {min_price}"),
        ("Most Common Rating", top_rating),
    ]

    positions = [
        ("A4", "B6"),
        ("D4", "E6"),
        ("G4", "H6"),
        ("A8", "B10"),
        ("D8", "E10"),
    ]

    # ======================================
    # KPI CARDS
    # ======================================

    for (title, value), (start, end) in zip(metrics, positions):

        dashboard.merge_cells(f"{start}:{end}")

        cell = dashboard[start]

        cell.value = f"{title}\n\n{value}"

        cell.fill = CARD_FILL

        cell.font = VALUE_FONT

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

    # ======================================
    # COLUMN WIDTHS
    # ======================================

    for column in ["A", "B", "D", "E", "G", "H"]:
        dashboard.column_dimensions[column].width = 22

    # Spacer columns
    dashboard.column_dimensions["C"].width = 6
    dashboard.column_dimensions["F"].width = 6

    # ======================================
    # ROW HEIGHTS
    # ======================================

    for row in [4, 5, 6, 8, 9, 10]:
        dashboard.row_dimensions[row].height = 35

    logger.info(
        "Dashboard created successfully"
    )